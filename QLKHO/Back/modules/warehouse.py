import sqlite3
import os
import traceback
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import qrcode
import tempfile
from copy import copy

try:
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    HAS_ADVANCED_ANCHOR = True
except ImportError:
    HAS_ADVANCED_ANCHOR = False

warehouse_bp = Blueprint('warehouse', __name__)

def get_db_connection():
    db_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data.db'))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

# =========================================================
# PHẦN 1: QUẢN LÝ ĐƠN HÀNG (TẠO / SỬA / LẤY THÔNG TIN)
# =========================================================

@warehouse_bp.route('/get-customers', methods=['GET'])
def get_customers():
    """API lấy danh sách đối tác khách hàng cho form tạo đơn"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Ma_KH, Ten_KH FROM CUSTOMERS")
        customers = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'customers': customers}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@warehouse_bp.route('/save-order', methods=['POST'])
def save_order():
    """API Lưu Đơn hàng (Lưu tên tự do, không còn vướng khóa ngoại)"""
    try:
        data = request.get_json(force=True, silent=True) or {}
        
        ma_dh = str(data.get('ma_dh', '')).strip().upper()
        ma_kh = str(data.get('ma_kh', '')).strip()
        ngay_du_kien = str(data.get('ngay_du_kien', '')).strip()
        nguoi_tao = str(data.get('user_id', '')).strip() # Gõ gì lưu nấy
        ngay_tao = str(data.get('ngay_tao', '')).strip()
        chi_tiet_po = data.get('chi_tiet_po', [])

        missing = []
        if not ma_dh or ma_dh == 'NONE': missing.append("Mã Đơn Hàng")
        if not ma_kh or ma_kh == 'NONE': missing.append("Đối Tác")
        if not ngay_du_kien or ngay_du_kien == 'NONE': missing.append("Ngày Giao")
        if not nguoi_tao: missing.append("Người tạo đơn")
        if not ngay_tao: missing.append("Ngày tạo đơn")
        if not chi_tiet_po: missing.append("Chi tiết PO")

        if missing:
            return jsonify({'success': False, 'message': f'Thiếu thông tin: {", ".join(missing)}'}), 400

        if len(ngay_tao) == 10:
            ngay_tao += datetime.now().strftime(' %H:%M:%S')

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT Trang_thai_DH FROM ORDERS WHERE Ma_DH = ?", (ma_dh,))
        existing_order = cursor.fetchone()

        if existing_order:
            if existing_order['Trang_thai_DH'] != 'Chờ xuất kho':
                conn.close()
                return jsonify({'success': False, 'message': f'Đơn hàng đang "{existing_order["Trang_thai_DH"]}", không thể sửa!'}), 400
            
            # Lưu trực tiếp text nguoi_tao vào cột User_ID
            cursor.execute("UPDATE ORDERS SET Ma_KH = ?, Ngay_giao = ?, User_ID = ?, Ngay_tao = ? WHERE Ma_DH = ?", 
                           (ma_kh, ngay_du_kien, nguoi_tao, ngay_tao, ma_dh))
            cursor.execute("DELETE FROM ORDER_DETAILS WHERE Ma_DH = ?", (ma_dh,))
        else:
            # Lưu trực tiếp text nguoi_tao vào cột User_ID
            cursor.execute("""
                INSERT INTO ORDERS (Ma_DH, Ma_KH, User_ID, Ngay_tao, Ngay_giao, Trang_thai_DH)
                VALUES (?, ?, ?, ?, ?, 'Chờ xuất kho')
            """, (ma_dh, ma_kh, nguoi_tao, ngay_tao, ngay_du_kien))

        for po in chi_tiet_po:
            cursor.execute("""
                INSERT INTO ORDER_DETAILS (Detail_ID, Ma_DH, MSP, So_luong)
                VALUES (?, ?, ?, ?)
            """, (str(po['ma_po']).strip().upper(), ma_dh, po['msp'], po['so_luong_pcs']))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Lưu đơn hàng thành công!'}), 200

    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': f'Lỗi: Mã PO đã tồn tại hoặc bị trùng lặp!'}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500


@warehouse_bp.route('/get-order/<order_id>', methods=['GET'])
def get_order(order_id):
    """API Lấy thông tin đơn hàng cũ để load lên form Sửa"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ORDERS WHERE Ma_DH = ?", (order_id,))
        order = cursor.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': 'Không tìm thấy đơn hàng!'}), 404
            
        cursor.execute("SELECT * FROM ORDER_DETAILS WHERE Ma_DH = ?", (order_id,))
        details = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({
            'success': True, 
            'order': dict(order),
            'details': details
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =========================================================
# PHẦN 2: CÁC API ĐÓNG THÙNG VÀ IN TEM THÙNG 
# =========================================================

def copy_row_block_thung(ws, src_start, src_end, tgt_start):
    row_count = src_end - src_start + 1
    row_offset = tgt_start - src_start
    
    for r in range(row_count):
        s_r = src_start + r
        t_r = tgt_start + r
        if s_r in ws.row_dimensions and ws.row_dimensions[s_r].height:
            ws.row_dimensions[t_r].height = ws.row_dimensions[s_r].height
        else:
            ws.row_dimensions[t_r].height = 20

        for col in range(1, 16):
            s_cell = ws.cell(row=s_r, column=col)
            t_cell = ws.cell(row=t_r, column=col)
            if isinstance(s_cell, MergedCell) or isinstance(t_cell, MergedCell):
                continue
            try:
                t_cell.value = s_cell.value
                if s_cell.has_style:
                    t_cell.font = copy(s_cell.font)
                    t_cell.border = copy(s_cell.border)
                    t_cell.fill = copy(s_cell.fill)
                    t_cell.number_format = s_cell.number_format
                    t_cell.alignment = copy(s_cell.alignment)
            except Exception:
                pass

    for m_range in list(ws.merged_cells.ranges):
        if m_range.min_row >= src_start and m_range.max_row <= src_end:
            n_min_r = m_range.min_row + row_offset
            n_max_r = m_range.max_row + row_offset
            col_min_letter = get_column_letter(m_range.min_col)
            col_max_letter = get_column_letter(m_range.max_col)
            try:
                ws.merge_cells(f"{col_min_letter}{n_min_r}:{col_max_letter}{n_max_r}")
            except Exception:
                pass


@warehouse_bp.route('/verify-bag', methods=['POST'])
def verify_bag():
    data = request.get_json()
    bag_id = data.get('bag_id', '').strip()

    if not bag_id: return jsonify({'success': False, 'message': 'Vui lòng quét mã túi!'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT b.*, p.Ten_SP, p.Quy_cach_Thung 
            FROM BAGS b
            JOIN PRODUCTS p ON b.MSP = p.MSP
            WHERE b.Bag_ID = ?
        ''', (bag_id,))
        bag = cursor.fetchone()
        conn.close()

        if not bag:
            return jsonify({'success': False, 'message': f'Mã {bag_id} không tồn tại!'}), 404
        if bag['Box_ID'] is not None:
            return jsonify({'success': False, 'message': f'Túi {bag_id} ĐÃ NẰM TRONG THÙNG {bag["Box_ID"]} rồi!'}), 400
        
        if bag['Trang_thai'] not in ['Đạt chuẩn', 'Bình thường']:
            return jsonify({'success': False, 'message': f'Túi {bag_id} chưa qua QC hoặc đang lỗi ({bag["Trang_thai"]})!'}), 400

        return jsonify({'success': True, 'bag': dict(bag)}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500


@warehouse_bp.route('/pack-box', methods=['POST'])
def pack_box():
    data = request.get_json()
    bag_ids = data.get('bags', [])
    user_id = data.get('user_id', 'NTV-0006')

    if not bag_ids: return jsonify({'success': False, 'message': 'Không có túi nào để đóng thùng!'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        now = datetime.now()
        date_str = now.strftime('%y%m%d')
        
        cursor.execute("SELECT MAX(CAST(SUBSTR(Box_ID, 9) AS INTEGER)) FROM BOXES WHERE Box_ID LIKE ?", (f"T-{date_str}%",))
        max_seq = cursor.fetchone()[0]
        start_seq = (max_seq + 1) if max_seq else 1
        box_id = f"T-{date_str}{str(start_seq).zfill(2)}"

        cursor.execute("SELECT MSP FROM BAGS WHERE Bag_ID = ?", (bag_ids[0],))
        msp = cursor.fetchone()['MSP']

        qty_bags = len(bag_ids)
        time_pack = now.strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute("INSERT INTO BOXES (Box_ID, MSP, Ngay_dong_goi, Ngay_nhap_kho, So_luong_Tui_Thuc_Te, Trang_thai) VALUES (?, ?, ?, ?, ?, 'Tồn kho')",
                       (box_id, msp, time_pack, time_pack, qty_bags))
        
        trans_id = f"IN-{box_id}"
        cursor.execute("INSERT INTO INVENTORY_TRANSACTIONS (Trans_ID, Box_ID, Loai_giao_dich, Ma_Tham_Chieu, Thoi_gian) VALUES (?, ?, 'Nhập kho', NULL, ?)",
               (trans_id, box_id, time_pack))

        for b_id in bag_ids:
            cursor.execute("UPDATE BAGS SET Box_ID = ? WHERE Bag_ID = ?", (box_id, b_id))

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': f'Đóng thành công thùng {box_id}!', 'box_id': box_id}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500


@warehouse_bp.route('/print-box-labels', methods=['POST'])
def print_box_labels():
    data = request.get_json() or {}
    box_ids = data.get('box_ids', [])

    if not box_ids: return jsonify({'success': False, 'message': 'Vui lòng chọn thùng cần in!'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates', 'form_thung.xlsx'))
        temp_dir = tempfile.mkdtemp()
        
        if not os.path.exists(template_path):
            conn.close()
            return jsonify({'success': False, 'message': f'Không tìm thấy file mẫu "form_thung.xlsx"'}), 404

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        for idx, box_id in enumerate(box_ids):
            cursor.execute("SELECT * FROM BOXES WHERE Box_ID = ?", (box_id,))
            box = cursor.fetchone()
            if not box: continue

            msp = box['MSP']
            qty_bags = box['So_luong_Tui_Thuc_Te'] if box['So_luong_Tui_Thuc_Te'] is not None else 0

            cursor.execute("SELECT DISTINCT So_Lot, NSX FROM BAGS WHERE Box_ID = ?", (box_id,))
            bags_info = cursor.fetchall()
            
            lots_list = []
            dates_list = []
            for row in bags_info:
                if row['So_Lot'] and row['So_Lot'] not in lots_list: lots_list.append(row['So_Lot'])
                if row['NSX']:
                    try:
                        d_obj = datetime.strptime(str(row['NSX']).split(' ')[0], '%Y-%m-%d')
                        d_str = d_obj.strftime('%d-%m-%y')
                        if d_str not in dates_list: dates_list.append(d_str)
                    except Exception:
                        dates_list.append(str(row['NSX']))

            lot_str = ", ".join(lots_list) if lots_list else f"LOT-{box_id}"
            date_str = ", ".join(dates_list) if dates_list else datetime.now().strftime('%d-%m-%y')

            cursor.execute("SELECT * FROM PRODUCTS WHERE MSP = ?", (msp,))
            product_info = cursor.fetchone()

            if product_info:
                ten_sp = product_info['Ten_SP']
                keys = product_info.keys() if hasattr(product_info, 'keys') else []
                customer_name = product_info['NCC'] if ('NCC' in keys and product_info['NCC']) else "Đối tác tiêu chuẩn"
                so_luong_pcs_trong_tui = product_info['Quy_cach_Tui'] if ('Quy_cach_Tui' in keys and product_info['Quy_cach_Tui']) else 125
            else:
                ten_sp = msp
                customer_name = "Đối tác tiêu chuẩn"
                so_luong_pcs_trong_tui = 125

            total_pcs = qty_bags * so_luong_pcs_trong_tui

            grid_col = idx % 3        
            grid_row = idx // 3       
            
            top_row = 1 + (grid_row * 8)    
            left_col = 1 + (grid_col * 5)   

            if grid_row >= 3 and grid_col == 0:
                copy_row_block_thung(ws, 1, 8, top_row)

            customer_col_letter = get_column_letter(left_col + 1)
            part_no_col_letter  = get_column_letter(left_col + 3)
            qr_col_letter       = get_column_letter(left_col + 3)

            def set_val(cell_ref, val):
                try:
                    c = ws[cell_ref]
                    if not isinstance(c, MergedCell): c.value = val
                except Exception: pass

            set_val(f"{customer_col_letter}{top_row}", customer_name)
            set_val(f"{part_no_col_letter}{top_row}", msp)
            set_val(f"{customer_col_letter}{top_row + 1}", ten_sp)
            set_val(f"{customer_col_letter}{top_row + 2}", f"{qty_bags} Bags ({total_pcs} Pcs)")
            set_val(f"{customer_col_letter}{top_row + 3}", date_str)
            set_val(f"{customer_col_letter}{top_row + 4}", lot_str)
            set_val(f"{customer_col_letter}{top_row + 5}", box_id)
            
            qr = qrcode.QRCode(box_size=4, border=1)
            qr.add_data(box_id) 
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            
            qr_path = os.path.join(temp_dir, f"{box_id}.png")
            img.save(qr_path)
            
            xl_img = XLImage(qr_path)
            xl_img.width = 85
            xl_img.height = 85
            
            qr_row_idx_1based = top_row + 3
            
            if HAS_ADVANCED_ANCHOR:
                col_idx_0based = left_col + 2  
                row_idx_0based = qr_row_idx_1based - 1
                marker = AnchorMarker(col=col_idx_0based, colOff=0, row=row_idx_0based, rowOff=int(5 * 9525))
                size = XDRPositiveSize2D(int(85 * 9525), int(85 * 9525))
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(xl_img)
            else:
                ws.add_image(xl_img, f"{qr_col_letter}{qr_row_idx_1based}")

        conn.close()

        output_file_name = f"Tem_Thung_{box_ids[0]}.xlsx" if len(box_ids) == 1 else f"In_Gop_Tem_Thung_{len(box_ids)}_Thung.xlsx"
        output_file_path = os.path.join(temp_dir, output_file_name)
        wb.save(output_file_path)

        return send_file(output_file_path, as_attachment=True, download_name=output_file_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500

@warehouse_bp.route('/print-box-label/<box_id>', methods=['GET'])
def print_box_label(box_id):
    try:
        from flask import current_app
        with current_app.test_request_context(json={'box_ids': [box_id]}):
            return print_box_labels()
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =========================================================
# PHẦN 3: TRẠM QUÉT XUẤT KHO (ĐI ĐƠN)
# =========================================================

@warehouse_bp.route('/get-pending-orders', methods=['GET'])
def get_pending_orders():
    """Lấy danh sách các đơn hàng đang chờ xuất kho"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT o.Ma_DH, o.Ngay_tao, o.Ngay_giao, c.Ten_KH, o.Trang_thai_DH
            FROM ORDERS o
            JOIN CUSTOMERS c ON o.Ma_KH = c.Ma_KH
            WHERE o.Trang_thai_DH = 'Chờ xuất kho'
            ORDER BY o.Ngay_tao ASC
        """)
        orders = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'orders': orders}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@warehouse_bp.route('/order-export-status/<ma_dh>', methods=['GET'])
def get_order_export_status(ma_dh):
    """Lấy chi tiết tiến độ quét xuất kho của từng PO trong 1 Đơn hàng"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                od.Detail_ID, 
                od.MSP, 
                p.Ten_SP, 
                od.So_luong AS Total_Req,
                COALESCE((
                    SELECT SUM(b.So_luong_Tui_Thuc_Te * p2.Quy_cach_Tui)
                    FROM BOXES b 
                    JOIN PRODUCTS p2 ON b.MSP = p2.MSP
                    WHERE b.Detail_ID = od.Detail_ID AND b.Trang_thai = 'Đã xuất'
                ), 0) AS Total_Scanned
            FROM ORDER_DETAILS od
            JOIN PRODUCTS p ON od.MSP = p.MSP
            WHERE od.Ma_DH = ?
        """, (ma_dh,))
        
        details = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'details': details}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@warehouse_bp.route('/scan-export', methods=['POST'])
def scan_export():
    """Xử lý khi thủ kho tít 1 mã thùng để xuất kho"""
    data = request.get_json(force=True, silent=True) or {}
    ma_dh = data.get('ma_dh', '').strip()
    box_id = data.get('box_id', '').strip().upper()
    user_id = data.get('user_id', 'NTV-0006')

    if not ma_dh or not box_id:
        return jsonify({'success': False, 'message': 'Thiếu mã đơn hàng hoặc mã thùng!'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT b.*, p.Quy_cach_Tui FROM BOXES b JOIN PRODUCTS p ON b.MSP = p.MSP WHERE b.Box_ID = ?", (box_id,))
        box = cursor.fetchone()

        if not box:
            return jsonify({'success': False, 'message': f'Thùng {box_id} không tồn tại!'}), 404
        if box['Trang_thai'] != 'Tồn kho':
            return jsonify({'success': False, 'message': f'Thùng {box_id} đang ở trạng thái "{box["Trang_thai"]}"!'}), 400
        if box['Detail_ID'] is not None:
            return jsonify({'success': False, 'message': f'Thùng {box_id} đã được xuất cho PO {box["Detail_ID"]}!'}), 400

        box_msp = box['MSP']
        box_pcs = (box['So_luong_Tui_Thuc_Te'] or 0) * (box['Quy_cach_Tui'] or 125)

        cursor.execute("""
            SELECT od.Detail_ID, od.So_luong AS Total_Req,
                COALESCE((
                    SELECT SUM(b2.So_luong_Tui_Thuc_Te * p2.Quy_cach_Tui)
                    FROM BOXES b2 JOIN PRODUCTS p2 ON b2.MSP = p2.MSP
                    WHERE b2.Detail_ID = od.Detail_ID AND b2.Trang_thai = 'Đã xuất'
                ), 0) AS Total_Scanned
            FROM ORDER_DETAILS od
            WHERE od.Ma_DH = ? AND od.MSP = ?
        """, (ma_dh, box_msp))
        
        pos = cursor.fetchall()
        
        if not pos:
            return jsonify({'success': False, 'message': f'Đơn hàng này KHÔNG YÊU CẦU sản phẩm {box_msp}!'}), 400

        target_po = None
        for po in pos:
            if po['Total_Scanned'] < po['Total_Req']:
                target_po = po['Detail_ID']
                break
        
        if not target_po:
            target_po = pos[0]['Detail_ID']

        cursor.execute("UPDATE BOXES SET Trang_thai = 'Đã xuất', Detail_ID = ? WHERE Box_ID = ?", (target_po, box_id))

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        trans_id = f"OUT-{box_id}-{ma_dh}"
        cursor.execute("""
              INSERT INTO INVENTORY_TRANSACTIONS (Trans_ID, Box_ID, Loai_giao_dich, Ma_Tham_Chieu, Thoi_gian)
               VALUES (?, ?, 'Xuất kho', ?, ?)
        """, (trans_id, box_id, ma_dh, now_str))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True, 
            'message': f'Đã xuất {box_pcs} Pcs sản phẩm {box_msp} vào PO: {target_po}!'
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500


@warehouse_bp.route('/complete-order', methods=['POST'])
def complete_order():
    """Chốt đơn hàng khi đã quét xong (CÓ KIỂM TRA SỐ LƯỢNG)"""
    data = request.get_json(force=True, silent=True) or {}
    ma_dh = data.get('ma_dh', '').strip()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                od.Detail_ID, 
                od.So_luong AS Total_Req,
                COALESCE((
                    SELECT SUM(b.So_luong_Tui_Thuc_Te * p2.Quy_cach_Tui)
                    FROM BOXES b 
                    JOIN PRODUCTS p2 ON b.MSP = p2.MSP
                    WHERE b.Detail_ID = od.Detail_ID AND b.Trang_thai = 'Đã xuất'
                ), 0) AS Total_Scanned
            FROM ORDER_DETAILS od
            WHERE od.Ma_DH = ?
        """, (ma_dh,))
        
        pos = cursor.fetchall()
        
        for po in pos:
            if po['Total_Scanned'] < po['Total_Req']:
                conn.close()
                return jsonify({
                    'success': False, 
                    'message': f'PO "{po["Detail_ID"]}" chưa đủ hàng ({po["Total_Scanned"]}/{po["Total_Req"]} Pcs). Không thể chốt! Nếu muốn đi thiếu, hãy quay lại sửa đơn hàng.'
                }), 400

        cursor.execute("UPDATE ORDERS SET Trang_thai_DH = 'Đã giao' WHERE Ma_DH = ?", (ma_dh,))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Đã chốt hoàn thành Đơn Hàng!'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@warehouse_bp.route('/get-all-orders', methods=['GET'])
def get_all_orders():
    """Lấy toàn bộ danh sách đơn hàng phục vụ cho trang Lịch sử / Quản lý đơn"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT o.Ma_DH, o.Ngay_tao, o.Ngay_giao, c.Ten_KH, o.Trang_thai_DH, o.User_ID
            FROM ORDERS o
            JOIN CUSTOMERS c ON o.Ma_KH = c.Ma_KH
            ORDER BY o.Ngay_tao DESC
        """)
        orders = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'orders': orders}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500    
    
@warehouse_bp.route('/delete-order/<ma_dh>', methods=['DELETE', 'POST'])
def delete_order(ma_dh):
    """Xóa đơn hàng (chỉ khi đang Chờ xuất kho) và tự động giải phóng các thùng đã lỡ quét"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Kiểm tra trạng thái đơn hàng
        cursor.execute("SELECT Trang_thai_DH FROM ORDERS WHERE Ma_DH = ?", (ma_dh,))
        order = cursor.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': 'Không tìm thấy đơn hàng!'}), 404
            
        if order['Trang_thai_DH'] != 'Chờ xuất kho':
            conn.close()
            return jsonify({'success': False, 'message': f'Đơn hàng đang ở trạng thái "{order["Trang_thai_DH"]}", không thể xóa!'}), 400
            
        # 2. Lấy danh sách Detail_ID của đơn hàng này để giải phóng các thùng đã quét (nếu có)
        cursor.execute("SELECT Detail_ID FROM ORDER_DETAILS WHERE Ma_DH = ?", (ma_dh,))
        details = cursor.fetchall()
        detail_ids = [d['Detail_ID'] for d in details]
        
        if detail_ids:
            placeholders = ','.join(['?'] * len(detail_ids))
            # Reset trạng thái các thùng đã gán PO này về 'Tồn kho' và xóa Detail_ID
            cursor.execute(f"UPDATE BOXES SET Trang_thai = 'Tồn kho', Detail_ID = NULL WHERE Detail_ID IN ({placeholders})", detail_ids)
            
        # 3. Xóa chi tiết PO và Đơn hàng chính
        cursor.execute("DELETE FROM ORDER_DETAILS WHERE Ma_DH = ?", (ma_dh,))
        cursor.execute("DELETE FROM ORDERS WHERE Ma_DH = ?", (ma_dh,))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'Đã xóa thành công đơn hàng {ma_dh} và giải phóng các thùng liên quan!'}), 200
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500