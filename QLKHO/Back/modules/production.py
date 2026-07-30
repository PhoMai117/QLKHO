import sqlite3
import os
import traceback
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import qrcode
import tempfile
from copy import copy

# Thư viện căn chỉnh vị trí ảnh chính xác từng pixel trong Excel
try:
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    HAS_ADVANCED_ANCHOR = True
except ImportError:
    HAS_ADVANCED_ANCHOR = False

production_bp = Blueprint('production', __name__)

def get_db_connection():
    db_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data.db'))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

# =========================================================
# 1. API LẤY DANH SÁCH SẢN PHẨM
# =========================================================
@production_bp.route('/get-products', methods=['GET'])
def get_products():
    """API lấy danh sách sản phẩm cho dropdown ở giao diện xưởng"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT MSP, Ten_SP, Quy_cach_Tui FROM PRODUCTS")
        products = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'products': products}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def copy_row_block(ws, src_start, src_end, tgt_start):
    """Sao chép block hàng mẫu nếu số tem vượt quá 16 tem (trôi sang trang sau)"""
    row_count = src_end - src_start + 1
    row_offset = tgt_start - src_start
    
    for r in range(row_count):
        s_r = src_start + r
        t_r = tgt_start + r
        if s_r in ws.row_dimensions and ws.row_dimensions[s_r].height:
            ws.row_dimensions[t_r].height = ws.row_dimensions[s_r].height
        else:
            ws.row_dimensions[t_r].height = 20

        for col in range(1, 21):
            s_cell = ws.cell(row=s_r, column=col)
            t_cell = ws.cell(row=t_r, column=col)
            if isinstance(s_cell, MergedCell) or isinstance(t_cell, MergedCell):
                continue
            try:
                t_cell.value = s_cell.value
                if s_cell.has_style:
                    t_cell.font = copy(s_cell.font)
                    t_cell.border = copy(s_cell.border)
                    t_cell.fill = copy(s_cell.fill)
                    t_cell.number_format = s_cell.number_format
                    t_cell.alignment = copy(s_cell.alignment)
            except Exception:
                pass

    for m_range in list(ws.merged_cells.ranges):
        if m_range.min_row >= src_start and m_range.max_row <= src_end:
            n_min_r = m_range.min_row + row_offset
            n_max_r = m_range.max_row + row_offset
            col_min_letter = get_column_letter(m_range.min_col)
            col_max_letter = get_column_letter(m_range.max_col)
            try:
                ws.merge_cells(f"{col_min_letter}{n_min_r}:{col_max_letter}{n_max_r}")
            except Exception:
                pass


# =========================================================
# 2. API IN TEM TÚI (CĂN GIỮA MÃ QR CODE TRONG VÙNG GỘP)
# =========================================================
@production_bp.route('/generate-bags', methods=['POST'])
def generate_bags():
    data = request.get_json()
    msp = data.get('msp')
    nsx = data.get('nsx') 
    lot_no = data.get('lot_no')
    so_luong_tem = int(data.get('so_luong_tem', 1))

    if not msp or not nsx or not lot_no:
        return jsonify({'success': False, 'message': 'Vui lòng điền đủ thông tin!'}), 400

    try:
        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates', 'form_tem.xlsx'))
        
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'message': f'Không tìm thấy file "form_tem.xlsx" tại: {template_path}'}), 404

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM PRODUCTS WHERE MSP = ?", (msp,))
        sp = cursor.fetchone()
        if not sp:
            conn.close()
            return jsonify({'success': False, 'message': 'Sản phẩm không tồn tại!'}), 404
        
        ten_sp = sp['Ten_SP']
        quy_cach_tui = sp['Quy_cach_Tui']
        keys = sp.keys() if hasattr(sp, 'keys') else []
        
        # --- SỬA LỖI TÊN CỘT NCC ---
        customer_name = sp['NCC'] if ('NCC' in keys and sp['NCC']) else "Đối tác tiêu chuẩn"

        # --- SỬA LỖI ĐỊNH DẠNG NGÀY DD-MM-YY ---
        date_obj = datetime.strptime(nsx, '%Y-%m-%d')
        date_str = date_obj.strftime('%y%m%d')      # Dùng để sinh mã Bag_ID (giữ nguyên)
        nsx_in_tem = date_obj.strftime('%d-%m-%y')  # Dùng để in trực tiếp lên tem

        prefix = f"B-{date_str}"
        cursor.execute("SELECT Bag_ID FROM BAGS WHERE Bag_ID LIKE ?", (f"{prefix}%",))
        existing_bags = cursor.fetchall()
        
        max_seq = 0
        for row in existing_bags:
            try:
                seq_num = int(row['Bag_ID'][len(prefix):])
                if seq_num > max_seq: max_seq = seq_num
            except ValueError:
                continue
        start_seq = max_seq + 1

        # Mở file mẫu Excel ma trận 4x4
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        temp_dir = tempfile.mkdtemp()

        # ĐIỀN LẦN LƯỢT VÀO CÁC Ô TRONG MA TRẬN 4 HÀNG x 4 CỘT
        for i in range(so_luong_tem):
            current_seq = start_seq + i
            bag_id = f"{prefix}{str(current_seq).zfill(3)}"
            
            cursor.execute("INSERT INTO BAGS (Bag_ID, MSP, NSX, So_Lot, Trang_thai) VALUES (?, ?, ?, ?, 'Mới in')", 
                           (bag_id, msp, nsx, lot_no))

            # Chỉ số cột (0..3) và hàng (0..3)
            grid_col = i % 4        
            grid_row = i // 4       
            
            top_row = 1 + (grid_row * 8)    # Mỗi tem cao 7 dòng + 1 dòng cách = 8
            left_col = 1 + (grid_col * 5)   # Mỗi tem rộng 4 cột + 1 cột cách = 5

            # Copy khung mẫu nếu vượt quá 16 tem
            if grid_row >= 4 and grid_col == 0:
                copy_row_block(ws, 1, 8, top_row)

            # Tọa độ cột chứa chữ & QR
            customer_col_letter = get_column_letter(left_col + 1) # B, G, L, Q
            part_no_col_letter  = get_column_letter(left_col + 3) # D, I, N, S
            qr_col_letter       = get_column_letter(left_col + 2) # C, H, M, R

            # Hàm gán giá trị an toàn tránh ô gộp phụ
            def set_cell_val(cell_ref, val):
                try:
                    cell = ws[cell_ref]
                    if not isinstance(cell, MergedCell):
                        cell.value = val
                except Exception:
                    pass

            # 1. Điền dữ liệu văn bản
            set_cell_val(f"{customer_col_letter}{top_row}", customer_name)             # Customer
            set_cell_val(f"{part_no_col_letter}{top_row}", msp)                        # Part No (MSP)
            set_cell_val(f"{customer_col_letter}{top_row + 1}", ten_sp)                 # Part Name
            set_cell_val(f"{customer_col_letter}{top_row + 2}", nsx_in_tem)             # --- SỬA Ở ĐÂY: In NSX theo DD-MM-YY ---
            set_cell_val(f"{customer_col_letter}{top_row + 3}", f"{quy_cach_tui} Pcs")  # Quantity
            set_cell_val(f"{customer_col_letter}{top_row + 4}", lot_no)                 # Lot No

            # 2. Tạo hình QR Code
            qr = qrcode.QRCode(box_size=3, border=1)
            qr.add_data(bag_id)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            qr_path = os.path.join(temp_dir, f"{bag_id}.png")
            img.save(qr_path)
            
            # 3. Chèn QR Code và CĂN GIỮA VÀO VÙNG Ô GỘP C4
            xl_img = XLImage(qr_path)
            img_w_px = 75
            img_h_px = 75
            xl_img.width = img_w_px
            xl_img.height = img_h_px
            
            qr_row_idx_1based = top_row + 3   

            col_offset_px = 45  
            row_offset_px = 8   

            if HAS_ADVANCED_ANCHOR:
                col_idx_0based = left_col + 1  
                row_idx_0based = qr_row_idx_1based - 1

                marker = AnchorMarker(
                    col=col_idx_0based, 
                    colOff=int(col_offset_px * 9525),  
                    row=row_idx_0based, 
                    rowOff=int(row_offset_px * 9525)
                )
                size = XDRPositiveSize2D(
                    int(img_w_px * 9525), 
                    int(img_h_px * 9525)
                )
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(xl_img)
            else:
                qr_cell_ref = f"{qr_col_letter}{qr_row_idx_1based}"
                ws.add_image(xl_img, qr_cell_ref)

        conn.commit()
        conn.close()

        output_file_path = os.path.join(temp_dir, f"In_Tem_{date_str}_{msp}.xlsx")
        wb.save(output_file_path)

        return send_file(output_file_path, as_attachment=True, download_name=f"In_Tem_{date_str}_{msp}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print("\n=== LỖI IN TEM TÚI ===")
        traceback.print_exc()
        print("======================\n")
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: Xem chi tiết trong Terminal.'}), 500

# =========================================================
# 3. API TRẠM QC KIỂM TRA & DUYỆT TÚI
# =========================================================
@production_bp.route('/verify-bag', methods=['POST'])
def verify_bag():
    data = request.get_json()
    bag_id = data.get('bag_id', '').strip()
    if not bag_id: return jsonify({'success': False, 'message': 'Vui lòng quét mã túi!'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM BAGS WHERE Bag_ID = ?", (bag_id,))
        bag = cursor.fetchone()

        if not bag:
            conn.close()
            return jsonify({'success': False, 'message': f'Mã túi {bag_id} KHÔNG TỒN TẠI!'}), 404

        tt = bag['Trang_thai']
        if tt == 'Bình thường':
            conn.close()
            return jsonify({'success': False, 'message': f'Túi {bag_id} đã được QC trước đó rồi!'}), 400
        if tt != 'Mới in':
            conn.close()
            return jsonify({'success': False, 'message': f'Túi {bag_id} đang ở trạng thái "{tt}", không thể QC!'}), 400

        # ĐÃ SỬA: Cập nhật thành 'Bình thường'
        cursor.execute("UPDATE BAGS SET Trang_thai = 'Bình thường' WHERE Bag_ID = ?", (bag_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'QC THÀNH CÔNG túi {bag_id}!'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500